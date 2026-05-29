#!/usr/bin/env python3
"""
FrameMaker XLIFF -> OpenAI API -> Translated XLIFF
"""

import os, sys, re, json, time, copy, argparse, logging
import base64, gzip, html
from pathlib import Path
from lxml import etree
from openai import OpenAI
from image_ocr_translator import process_xlf_references

# ─────────────────────────────────────────────────────────────────────────────
# OpenAI client 
# ─────────────────────────────────────────────────────────────────────────────

import os
from dotenv import load_dotenv

load_dotenv()

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY environment variable not set")
client = OpenAI(api_key=OPENAI_API_KEY)

MODEL       = "gpt-4o"
MAX_TOKENS  = 8096
BATCH_SIZE  = 40
BATCH_DELAY = 0.5

# ─────────────────────────────────────────────────────────────────────────────
# SUPPORTED LANGUAGES 
# ─────────────────────────────────────────────────────────────────────────────
LANGUAGES = {
    "zh-CN": "Simplified Chinese (简体中文)",
    "zh-TW": "Traditional Chinese (繁體中文)",
    "ja":    "Japanese (日本語)",
    "ko":    "Korean (한국어)",
    "de":    "German (Deutsch)",
    "fr":    "French (Français)",
    "es":    "Spanish (Español)",
    "ar":    "Arabic (العربية)",
    "pt":    "Portuguese (Português)",
    "it":    "Italian (Italiano)",
    "vi":    "Vietnamese (Tiếng Việt)",
    "nl":    "Dutch (Nederlands)",
    "pl":    "Polish (Polski)",
    "ru":    "Russian (Русский)",
    "tr":    "Turkish (Türkçe)",
    "sv":    "Swedish (Svenska)",
    "da":    "Danish (Dansk)",
    "fi":    "Finnish (Suomi)",
    "nb":    "Norwegian (Norsk Bokmål)",
    "cs":    "Czech (Čeština)",
}

FM_LANG = {k: k for k in LANGUAGES}


def select_language_interactive():
    lang_items = list(LANGUAGES.items()) 

    print("\n" + "═" * 62)
    print("  XLIFF TRANSLATOR — Language Selection")
    print("═" * 62)
    for i, (code, label) in enumerate(lang_items, 1):
        print(f"  {i:>2}.  {code:<8}  {label}")
    print("═" * 62)
    print("  Enter the NUMBER (e.g. 5) or CODE (e.g. de)")
    print("═" * 62)

    while True:
        raw = input("\n  Your choice: ").strip()

        if raw.isdigit():
            idx = int(raw) - 1
            if 0 <= idx < len(lang_items):
                code, label = lang_items[idx]
                print(f"\n  ✓  Selected: {label}  [{code}]\n")
                return code
            else:
                print(f"  ✗  Please enter a number between 1 and {len(lang_items)}.")
                continue

        if raw in LANGUAGES:
            print(f"\n  ✓  Selected: {LANGUAGES[raw]}  [{raw}]\n")
            return raw

        matches = [(c, l) for c, l in lang_items if raw.lower() in l.lower()]
        if len(matches) == 1:
            code, label = matches[0]
            print(f"\n  ✓  Matched: {label}  [{code}]\n")
            return code
        elif len(matches) > 1:
            print(f"  ✗  Ambiguous — matches: {', '.join(c for c,_ in matches)}. Be more specific.")
        else:
            print(f"  ✗  '{raw}' not recognised. Try a number or a code like 'de', 'ja', 'vi'.")

def ask_path(prompt, must_exist=True, is_file=True):
    while True:
        raw = input(f"\n  {prompt}: ").strip().strip('"').strip("'")
        p = Path(raw).expanduser().resolve()
        if not must_exist:
            return p
        if is_file and p.is_file():
            return p
        if not is_file and p.is_dir():
            return p
        kind = "file" if is_file else "folder"
        print(f"  ✗  {kind} not found: {p}")

DO_NOT_TRANSLATE = {
    "SYSTEM OK","CLASS 100","CO2 AUTO CAL","SYS IN OTEMP","TSNSR1 ERR",
    "TSNSR2 ERR","CO2 SNSR ERR","O2 SNSR ERR","REPL O2 SNSR","REPL IR SNSR",
    "REPLACE HEPA","ADD WATER","DOOR IS OPEN","CO2 IS HIGH","CO2 IS LOW",
    "TEMP IS HIGH","TEMP IS LOW","O2 IS HIGH","O2 IS LOW","RH IS LOW",
    "IR AUTOZ ERR","TANK1 LOW","TANK2 LOW","TANK 1 and 2 LOW",
    "RJ-11","RS485","RS-485","RS232","USB","ESD",
}

SAFETY_STYLES = {
    "Warning","Caution","Note","Important","Danger","WarningTitle","CautionTitle",
    "NoteTitle","ImportantTitle","WarningBody","CautionBody","NoteBody","ImportantBody",
    "Admonition","AdmonitionTitle","Hazard","SafetyNote",
}

SAFETY_RE = re.compile(
    r"^\s*(Warning|Caution|Important|Note|Danger|WARNING|CAUTION|IMPORTANT)\b",
    re.IGNORECASE,
)

GLOSSARY = {
    "zh-CN": {
        "Operating Instructions": "操作说明",
        "Biological Safety Cabinet": "生物安全柜",
        "Water Jacket": "水套",   "Incubator": "培养箱",
        "HEPA Filter": "HEPA过滤器", "Control Panel": "控制面板",
        "Setpoint": "设定点", "Calibration": "校准",
        "Warning": "警告",         "Caution": "注意",
        "Important": "重要",       "Note": "备注",
    },
    "zh-TW": {
        "Operating Instructions": "操作說明",
        "Biological Safety Cabinet": "生物安全櫃",
        "Warning": "警告", "Caution": "注意",
        "Important": "重要", "Note": "備註",
    },
    "ja": {
        "Operating Instructions": "取扱説明書",
        "Biological Safety Cabinet": "バイオセーフティキャビネット",
        "Warning": "警告", "Caution": "注意",
        "Important": "重要", "Note": "注",
    },
    "de": {
        "Warning": "Warnung", "Caution": "Vorsicht",
        "Important": "Wichtig", "Note": "Hinweis",
        "Operating Instructions": "Bedienungsanleitung",
    },
    "fr": {
        "Warning": "Avertissement", "Caution": "Attention",
        "Important": "Important", "Note": "Remarque",
        "Operating Instructions": "Mode d'emploi",
    },
    "es": {
        "Warning": "Advertencia", "Caution": "Precaución",
        "Important": "Importante", "Note": "Nota",
        "Operating Instructions": "Instrucciones de funcionamiento",
    },
    "vi": {
        "Warning": "Cảnh báo", "Caution": "Thận trọng",
        "Important": "Quan trọng", "Note": "Lưu ý",
        "Operating Instructions": "Hướng dẫn vận hành",
        "Biological Safety Cabinet": "Tủ an toàn sinh học",
        "HEPA Filter": "Bộ lọc HEPA", "Calibration": "Hiệu chuẩn",
    },
    "ko": {
        "Warning": "경고", "Caution": "주의",
        "Important": "중요", "Note": "참고",
    },
    "pt": {
        "Warning": "Aviso", "Caution": "Cuidado",
        "Important": "Importante", "Note": "Nota",
    },
    "it": {
        "Warning": "Avvertenza", "Caution": "Attenzione",
        "Important": "Importante", "Note": "Nota",
    },
    "ru": {
        "Warning": "Предупреждение", "Caution": "Осторожно",
        "Important": "Важно", "Note": "Примечание",
    },
    "nl": {
        "Warning": "Waarschuwing", "Caution": "Let op",
        "Important": "Belangrijk", "Note": "Opmerking",
    },
    "pl": {
        "Warning": "Ostrzeżenie", "Caution": "Uwaga",
        "Important": "Ważne", "Note": "Uwaga",
    },
    "tr": {
        "Warning": "Uyarı", "Caution": "Dikkat",
        "Important": "Önemli", "Note": "Not",
    },
}

XML_NS   = "http://www.w3.org/XML/1998/namespace"
XML_LANG = f"{{{XML_NS}}}lang"
XML_SPC  = f"{{{XML_NS}}}space"

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("fm_translate")

def detect_ns(root):
    tag = root.tag
    if "{" in tag: return tag.split("}")[0].lstrip("{")
    return root.get("xmlns", "")

def Q(tag, ns): return f"{{{ns}}}{tag}" if ns else tag

def load_xliff(path):
    parser = etree.XMLParser(remove_blank_text=False, resolve_entities=False, recover=True)
    tree = etree.parse(path, parser)
    root = tree.getroot()
    ns = detect_ns(root)
    log.info(f"Loaded: {path}  namespace: {ns or '(none)'}")
    return tree, root, ns

def _style_from_group(tu, ns):
    p = tu.getparent()
    while p is not None:
        tag = p.tag.split("}")[-1] if "}" in p.tag else p.tag
        if tag == "group":
            rn = p.get("resname", "")
            if rn: return rn
        p = p.getparent()
    return ""

def _inner_text(el):
    if el is None: return ""
    return re.sub(r"<[^>]+>", "", etree.tostring(el, encoding="unicode", with_tail=False)).strip()

def extract_units(root, ns):
    units   = []
    tag_tu  = Q("trans-unit", ns)
    tag_src = Q("source",     ns)
    tag_seg = Q("seg-source", ns)
    tag_mrk = Q("mrk",        ns)

    for tu in root.iter(tag_tu):
        if tu.get("translate", "yes").lower() == "no": continue
        tu_id   = tu.get("id", f"tu_{len(units):04d}")
        resname = tu.get("resname", "")
        style   = _style_from_group(tu, ns) or resname
        src_el  = tu.find(tag_src)
        seg_el  = tu.find(tag_seg)

        if seg_el is not None:
            seg_mrks = [m for m in seg_el.iter(tag_mrk) if m.get("mtype") == "seg"]
            for mrk in seg_mrks:
                units.append({
                    "id":         f"{tu_id}::mrk::{mrk.get('mid', '')}",
                    "tu_id":      tu_id,
                    "mrk_mid":    mrk.get("mid", ""),
                    "element":    tu,
                    "seg_src_el": seg_el,
                    "source":     mrk.text or "",
                    "style":      style,
                    "restype":    tu.get("restype", ""),
                })
        else:
            text = _inner_text(src_el)
            units.append({
                "id":         tu_id,
                "tu_id":      tu_id,
                "mrk_mid":    None,
                "element":    tu,
                "seg_src_el": None,
                "source":     text,
                "style":      style,
                "restype":    tu.get("restype", ""),
            })

    n_tu = sum(1 for _ in root.iter(Q("trans-unit", ns)))
    log.info(f"Extracted {len(units)} segments from {n_tu} trans-units")
    return units

_TEMP_UNIT_RE = re.compile(
    r"^[\s]*[+\-]?\d+[\d.,]*\s*$"
    r"|^[\s]*[°℃℉]\s*[CF]?\s*$"
    r"|^[\s]*[CF]\s*$"
    r"|^[\s]*to\s*$"
    r"|^[\s]*[~–\-]\s*$"
    r"|^[\s]*[°%()±]\s*$",
    re.IGNORECASE,
)

_PAGE_REF_TAIL_RE = re.compile(
    r"\b(page|figure|fig|table|chapter|section|step|item|part)\s+$",
    re.IGNORECASE,
)

def _is_unit_fragment(u):
    s = u["source"].strip()
    if not s:
        return True
    if len(s) <= 5 and _TEMP_UNIT_RE.match(s):
        return True
    if len(s) <= 3 and re.match(r"^[°CF%()±\s]+$", s, re.IGNORECASE):
        return True
    return False

def _is_page_number_suffix(prev_source, curr_source):
    curr = curr_source.strip()
    if not re.match(r"^\d+$", curr):
        return False
    return bool(_PAGE_REF_TAIL_RE.search(prev_source))

def merge_units(units):
    if not units:
        return units

    def is_small_old(u):
        s = u["source"].strip()
        return (
            len(s) <= 3 or
            re.match(r"^[°CF%()]+$", s) or
            re.match(r"^\d+°?$", s)
        )

    merged = []
    buffer = None
    for u in units:
        if buffer is None:
            buffer = u
            continue
        if u["tu_id"] == buffer["tu_id"] and (is_small_old(u) or is_small_old(buffer)):
            buffer["source"] = buffer["source"].rstrip() + " " + u["source"].lstrip()
        else:
            merged.append(buffer)
            buffer = u
    if buffer:
        merged.append(buffer)

    result = []
    i = 0
    while i < len(merged):
        u = merged[i]
        if _is_unit_fragment(u) and result:
            prev = result[-1]
            prev["source"] = prev["source"].rstrip() + " " + u["source"].lstrip()
        elif i + 1 < len(merged) and _is_unit_fragment(merged[i + 1]):
            combined = u["source"]
            j = i + 1
            while j < len(merged) and _is_unit_fragment(merged[j]):
                combined = combined.rstrip() + " " + merged[j]["source"].lstrip()
                j += 1
            u = dict(u)
            u["source"] = combined
            result.append(u)
            i = j
            continue
        else:
            result.append(u)
        i += 1

    final = []
    for u in result:
        curr_src = u["source"]
        if (
            final
            and re.match(r"^\d+$", curr_src.strip())
            and _is_page_number_suffix(final[-1]["source"], curr_src)
        ):
            final[-1]["source"] = final[-1]["source"].rstrip() + " " + curr_src.strip()
        else:
            final.append(u)

    log.info(f"merge_units: {len(units)} → {len(final)} segments after merging")
    return final

def classify(unit):
    src = unit["source"].strip()

    if not src:
        return "skip"

    if re.search(r"\d", src) and re.search(r"[A-Za-z\u00b0\u2103\u2109]", src):
        return "body"

    if re.match(r"^[\d\s.\/%\u00b0\u00d7xX\u00b1~\u2264\u2265<>]+$", src):
        if re.match(r"^\d+$", src):
            return "body"
        return "skip"

    if re.match(r"^https?://|^www\.", src):
        return "skip"

    if len(src) <= 1 and not src.isalpha():
        return "skip"

    if src.upper() in {d.upper() for d in DO_NOT_TRANSLATE}:
        return "skip"

    style = unit.get("style", "")
    if style and any(s in style for s in SAFETY_STYLES):
        return "safety"

    if SAFETY_RE.match(src):
        return "safety"

    return "body"


SYS_PROMPT = """You are a professional technical translator for laboratory equipment manuals.
Translate the segments from English into {lang}.
Rules:
1. NEVER translate these -- return them verbatim: {dnt}
2. Return ONLY plain text values. No XML tags in your response.
3. Glossary (use these exact translations): {glossary}
4. Segments with [SAFETY] prefix are safety-critical -- translate with maximum fidelity.
5. For segments containing temperatures like "-20°C to +60°C" or "(-4°F to +140°F)",
   preserve the numeric values and unit symbols exactly; only translate surrounding words.
6. For segments that are pure numbers (e.g. "30", "25", "20") return them verbatim unchanged.
7. Respond with ONLY a JSON object: {{"id":"translation"}}. No markdown, no explanation."""


def build_sys(target_lang):
    lang  = LANGUAGES.get(target_lang, target_lang)
    dnt   = ", ".join(f'"{t}"' for t in sorted(DO_NOT_TRANSLATE)[:12])
    gdict = GLOSSARY.get(target_lang, {})
    gloss = "; ".join(f'"{en}"->"{tr}"' for en, tr in list(gdict.items())[:12])
    return SYS_PROMPT.format(lang=lang, dnt=dnt or "(none)", glossary=gloss or "(none)")


def translate_batch(batch, target_lang, sys_prompt, dry_run, model_to_use):
    if dry_run:
        return {u["id"]: f"[DRY RUN] {u['source'][:50]}" for u in batch}

    payload = {
        (f"[SAFETY]{u['id']}" if u.get("_class") == "safety" else u["id"]): u["source"]
        for u in batch
    }
    user_msg = (
        f"Translate {len(batch)} segments into "
        f"{LANGUAGES.get(target_lang, target_lang)}.\n"
        f"Return ONLY JSON.\n\n"
        f"{json.dumps(payload, ensure_ascii=False, indent=2)}"
    )

    for attempt in range(1, 4):
        try:
            response = client.chat.completions.create(
                model=model_to_use,
                max_tokens=MAX_TOKENS,
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user",   "content": user_msg},
                ],
                response_format={"type": "json_object"},
            )
            raw = response.choices[0].message.content.strip()

            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

            result = json.loads(raw)
            return {k.replace("[SAFETY]", ""): v for k, v in result.items()}

        except json.JSONDecodeError as e:
            log.warning(f"Attempt {attempt} JSON error: {e}")
            if attempt == 3:
                return {u["id"]: u["source"] for u in batch}
            time.sleep(2 ** attempt)

        except Exception as e:
            err = str(e).lower()
            wait = 30 * attempt if ("rate" in err or "429" in err) else 5
            log.warning(f"Attempt {attempt}: {e} -- wait {wait}s")
            if attempt == 3:
                return {u["id"]: u["source"] for u in batch}
            time.sleep(wait)

    return {u["id"]: u["source"] for u in batch}


def strip_seg_source(root, ns):
    tag_seg = Q("seg-source", ns)
    removed = 0
    for seg_el in root.findall(f".//{tag_seg}"):
        parent = seg_el.getparent()
        if parent is not None:
            prev = seg_el.getprevious()
            tail = seg_el.tail or ""
            if prev is not None:
                prev.tail = (prev.tail or "") + tail
            else:
                parent.text = (parent.text or "") + tail
            parent.remove(seg_el)
            removed += 1
    log.info(f"Stripped {removed} <seg-source> elements (XLIFF 1.2 schema fix)")
    return removed


def _inject_translation_into_source_clone(tgt, seg_el, mid_map, tag_mrk, tag_g):
    parts = []
    for child in seg_el:
        ctag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
        if ctag == "mrk" and child.get("mtype") == "seg":
            mid = child.get("mid", "")
            translated = mid_map.get(mid, child.text or "")
            parts.append(translated)
            if child.tail:
                parts.append(child.tail)
        else:
            inner_parts = []
            for m in child.iter(tag_mrk):
                if m.get("mtype") == "seg":
                    inner_parts.append(mid_map.get(m.get("mid", ""), m.text or ""))
            if inner_parts:
                parts.append(" ".join(inner_parts))
            if child.tail:
                parts.append(child.tail)

    translated_text = "".join(parts)

    children = list(tgt)

    if not children:
        tgt.text = translated_text
        return

    tgt_text_is_content = bool(tgt.text and tgt.text.strip())

    all_no_translate = all(
        c.get("translate", "yes").lower() == "no" for c in children
    )
    if all_no_translate:
        placed = False
        if tgt_text_is_content:
            tgt.text = translated_text
            placed = True
        if not placed:
            for child in children:
                if child.tail and child.tail.strip():
                    child.tail = translated_text
                    placed = True
                    break
        if not placed:
            children[-1].tail = translated_text
        return

    if tgt_text_is_content:
        tgt.text = translated_text
        for child in children:
            if child.get("translate", "yes").lower() == "no":
                child.text = None
                child.tail = None
        return

    tgt.text = None

    placed = False
    for i, child in enumerate(children):
        is_last = (i == len(children) - 1)
        translatable = child.get("translate", "yes").lower() != "no"

        if not placed:
            if child.tail and child.tail.strip():
                child.tail = translated_text
                placed = True
                for j in range(i + 1, len(children)):
                    c = children[j]
                    if c.get("translate", "yes").lower() == "no":
                        c.text = None
                    c.tail = None
                return
            elif translatable and child.text and child.text.strip():
                child.text = translated_text
                child.tail = None
                placed = True
                for j in range(i + 1, len(children)):
                    c = children[j]
                    if c.get("translate", "yes").lower() == "no":
                        c.text = None
                    c.tail = None
                return
            elif is_last:
                child.tail = translated_text
                placed = True
            else:
                if not translatable:
                    child.text = None
                child.tail = None
        else:
            if not translatable:
                child.text = None
            child.tail = None


def write_back(units, translations, ns, target_lang):
    tag_src    = Q("source",     ns)
    tag_seg    = Q("seg-source", ns)
    tag_target = Q("target",     ns)
    tag_mrk    = Q("mrk",        ns)
    tag_g      = Q("g",          ns)
    lang_code  = FM_LANG.get(target_lang, target_lang)

    from collections import defaultdict
    by_tu = defaultdict(list)
    for u in units:
        by_tu[u["tu_id"]].append(u)

    updated = 0
    for tu_id, tu_units in by_tu.items():
        tu_el  = tu_units[0]["element"]
        seg_el = tu_units[0]["seg_src_el"]
        src_el = tu_el.find(tag_src)

        for old in tu_el.findall(tag_target):
            tu_el.remove(old)

        if seg_el is not None:
            mid_map = {}
            for u in tu_units:
                if u["mrk_mid"] is not None:
                    t = translations.get(u["id"])
                    if t is not None:
                        mid_map[u["mrk_mid"]] = t

            if not mid_map:
                mid_map = {
                    u["mrk_mid"]: u["source"]
                    for u in tu_units if u["mrk_mid"] is not None
                }

            if src_el is not None:
                tgt = copy.deepcopy(src_el)
            else:
                tgt = etree.Element(tag_target)

            tgt.tag = tag_target
            if XML_LANG in tgt.attrib:
                del tgt.attrib[XML_LANG]
            if XML_SPC in tgt.attrib:
                del tgt.attrib[XML_SPC]
            tgt.set(XML_LANG, lang_code)
            tgt.set("state", "translated")
            tgt.tail = src_el.tail if src_el is not None else seg_el.tail

            _inject_translation_into_source_clone(tgt, seg_el, mid_map, tag_mrk, tag_g)

            ref_el = src_el if src_el is not None else seg_el
            ref_idx = list(tu_el).index(ref_el)
            tu_el.insert(ref_idx + 1, tgt)
            updated += 1

        else:
            u = tu_units[0]
            t = translations.get(u["id"])
            if t is None:
                continue

            if src_el is not None:
                tgt = copy.deepcopy(src_el)
                tgt.tag = tag_target
                if XML_LANG in tgt.attrib:
                    del tgt.attrib[XML_LANG]
                if XML_SPC in tgt.attrib:
                    del tgt.attrib[XML_SPC]
                tgt.set(XML_LANG, lang_code)
                tgt.set("state", "translated")
                tgt.tail = src_el.tail
                children = list(tgt)
                if not children:
                    tgt.text = t
                else:
                    tgt.text = None
                    placed = False
                    for child in children:
                        if not placed and child.tail and child.tail.strip():
                            child.tail = t
                            placed = True
                        elif not placed and child.get("translate","yes").lower() != "no" \
                                and child.text and child.text.strip():
                            child.text = t
                            child.tail = None
                            placed = True
                        else:
                            if child.get("translate","yes").lower() == "no":
                                child.text = None
                            child.tail = None
                    if not placed:
                        if children:
                            children[-1].tail = t
                        else:
                            tgt.text = t
                idx = list(tu_el).index(src_el)
                tu_el.insert(idx + 1, tgt)
            else:
                tgt = etree.Element(tag_target)
                tgt.set(XML_LANG, lang_code)
                tgt.set("state", "translated")
                tgt.text = t
                tu_el.append(tgt)

            updated += 1

    return updated

def validate_xml(tree):
    try:
        raw = etree.tostring(tree.getroot(), encoding="unicode")
        etree.fromstring(raw.encode("utf-8"))
        log.info("XML validation passed -- output is well-formed")
        return True
    except etree.XMLSyntaxError as e:
        log.error(f"XML validation FAILED: {e}")
        return False

def set_header_lang(root, ns, target_lang):
    lc = FM_LANG.get(target_lang, target_lang)
    for f in root.iter(Q("file", ns)):
        f.set("target-language", lc)
    log.info(f"Header target-language -> {lc}")

def save_xliff(tree, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    tree.write(path, encoding="utf-8", xml_declaration=True, pretty_print=True)
    log.info(f"Saved -> {path}")

def export_safety_review(units, translations, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.warning("openpyxl not installed -- skipping safety_review.xlsx")
        return
    rows = [(u["id"], u.get("style",""), u["source"], translations.get(u["id"],""))
            for u in units if u.get("_class") == "safety"]
    if not rows:
        return
    wb = Workbook(); ws = wb.active; ws.title = "Safety Review"
    ws.append(["ID","Style","English","Translation","Reviewer Notes"])
    hf    = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="1A3A5C")
    rfill = PatternFill("solid", fgColor="FAEEDA")
    wrap  = Alignment(wrap_text=True, vertical="top")
    for c in ws[1]:
        c.font = hf; c.fill = hfill
    for row in rows:
        ws.append(list(row) + [""])
        for c in ws[ws.max_row]:
            c.fill = rfill; c.alignment = wrap
    for col, w in zip("ABCDE", [18, 18, 52, 52, 18]):
        ws.column_dimensions[col].width = w
    wb.save(path)
    log.info(f"Safety review -> {path}  ({len(rows)} segments)")

def load_checkpoint(path):
    if Path(path).exists():
        with open(path, "r", encoding="utf-8") as f:
            d = json.load(f)
        log.info(f"Checkpoint: {len(d)} done")
        return d
    return {}

def save_checkpoint(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

_OB_RE = re.compile(
    r'(<(?:[A-Za-z_][\w\-]*:)?ImportObFile\b[^>]*>)'   
    r'([^<]+)'                                          
    r'(</(?:[A-Za-z_][\w\-]*:)?ImportObFile>)',         
    re.IGNORECASE,
)


def _basename_of_mif_value(raw: str) -> str:
    decoded = html.unescape(raw.strip())
    decoded = decoded.replace("<u>", "/").replace("<c>", "/")
    decoded = decoded.replace("\\", "/").replace(":", "/")
    parts = [p for p in decoded.split("/") if p]
    return parts[-1] if parts else decoded


def _rewrite_mif_blob(mif: str, path_mapping: dict) -> tuple:
    count = 0
    matched_samples: list = []
    unmatched_samples: list = []
    total_seen = 0

    def _replace(match: re.Match) -> str:
        nonlocal count, total_seen
        head, current, tail = match.group(1), match.group(2), match.group(3)
        total_seen += 1

        if current in path_mapping:
            count += 1
            new = path_mapping[current]
            if len(matched_samples) < 5:
                matched_samples.append((current, new, "exact"))
            return f"{head}{new}{tail}"

        bn = _basename_of_mif_value(current)
        if bn and bn in path_mapping:
            count += 1
            new = path_mapping[bn]
            if len(matched_samples) < 5:
                matched_samples.append((current, new, f"basename={bn!r}"))
            return f"{head}{new}{tail}"

        if len(unmatched_samples) < 5:
            unmatched_samples.append((current, bn))
        return match.group(0)

    new_mif = _OB_RE.sub(_replace, mif)

    log.info(f"  MIF blob: scanned {total_seen} <ImportObFile> entr(ies); rewrote {count}")
    for orig, new, how in matched_samples:
        log.info(f"    ✓ {orig!r}\n         → {new!r}  [{how}]")
    if unmatched_samples:
        log.warning(f"    {len(unmatched_samples)} sample(s) did NOT match — showing the values + their derived basenames:")
        for orig, bn in unmatched_samples:
            log.warning(f"    ✗ value={orig!r}  basename={bn!r}")
        log.warning(f"    Mapping keys available (first 5 of {len(path_mapping)}):")
        for k in list(path_mapping.keys())[:5]:
            log.warning(f"        {k!r}")

    return new_mif, count


_DI_RE = re.compile(
    r'(<(?:[A-Za-z_][\w\-]*:)?ImportObFileDI\b[^>]*>)'   
    r'([^<]+)'                                          
    r'(</(?:[A-Za-z_][\w\-]*:)?ImportObFileDI>)',         
    re.IGNORECASE,
)

def _to_mif_path(path_str: str) -> str:
    path_str = path_str.replace("\\", "/")
    parts = [p for p in path_str.split("/") if p not in ("", ".")]
    mif_parts = []
    for part in parts:
        if part == "..":
            mif_parts.append("<u>")
        else:
            mif_parts.append("<c>" + part)
    mif_path = "".join(mif_parts)
    return html.escape(mif_path)

def _reencode_mif_to_blob(mif: str, original_was_gzipped: bool) -> str:
    raw = mif.encode("utf-8", errors="replace")
    if original_was_gzipped:
        raw = gzip.compress(raw)
    return base64.b64encode(raw).decode("ascii")


def update_xlf_references(xlf_path, path_mapping):
    if not path_mapping:
        log.warning("update_xlf_references: empty mapping; nothing to do")
        return

    filename_to_new: dict[str, str] = {}
    filename_to_new_lower: dict[str, str] = {}
    for new_path in set(path_mapping.values()):
        bn = Path(new_path.replace("\\", "/")).name
        filename_to_new[bn] = new_path
        filename_to_new_lower[bn.lower()] = new_path

    log.info(f"update_xlf_references: {len(filename_to_new)} unique filename(s) to rewrite")
    for bn, new_path in filename_to_new.items():
        log.info(f"  • {bn}  →  {new_path}")

    path_mapping_lower = {k.lower(): v for k, v in path_mapping.items()}

    parser = etree.XMLParser(remove_blank_text=False, recover=True)
    tree   = etree.parse(str(xlf_path), parser)
    root   = tree.getroot()

    internal_el = None
    for elem in root.iter():
        if elem.tag.split("}")[-1] == "internal-file":
            internal_el = elem
            break

    if internal_el is None or not (internal_el.text and internal_el.text.strip()):
        log.warning("update_xlf_references: no <internal-file> element in XLF")
        return

    try:
        raw_b64    = internal_el.text.strip()
        compressed = base64.b64decode(raw_b64)
        was_gzip   = compressed[:2] == b"\x1f\x8b"
        mif_str = (
            gzip.decompress(compressed).decode("utf-8", errors="replace")
            if was_gzip
            else compressed.decode("utf-8", errors="replace")
        )
    except Exception as e:
        log.error(f"update_xlf_references: failed to decode <internal-file> ({e})")
        return

    rewrite_count = 0
    miss_samples: list = []
    
    result = []
    pos = 0

    # Match ImportObFileDI elements and their corresponding ImportObFile elements
    for di_match in _DI_RE.finditer(mif_str):
        di_head = di_match.group(1)
        di_content = di_match.group(2)
        di_tail = di_match.group(3)
        
        decoded = html.unescape(di_content.strip())
        converted = decoded.replace("<u>", "../").replace("<c>", "/").replace("..//", "../")
        basename = Path(converted).name
        
        new_path = (
            path_mapping.get(di_content) or
            path_mapping_lower.get(di_content.lower()) or
            path_mapping.get(basename) or
            path_mapping_lower.get(basename.lower()) or
            path_mapping.get(converted) or
            path_mapping_lower.get(converted.lower()) or
            filename_to_new.get(basename) or
            filename_to_new_lower.get(basename.lower())
        )
        
        if not new_path:
            if len(miss_samples) < 10:
                miss_samples.append(di_content)
            continue
            
        ob_match = _OB_RE.search(mif_str, di_match.end(), di_match.end() + 1000)
        if ob_match is None:
            continue
            
        ob_head = ob_match.group(1)
        ob_content = ob_match.group(2)
        ob_tail = ob_match.group(3)
        
        if ob_content.strip() == "2.0 internal inset":
            continue
            
        new_mif_path = _to_mif_path(new_path)
        
        # Append from last position up to the start of ImportObFileDI's content
        result.append(mif_str[pos : di_match.start(2)])
        # Append new ImportObFileDI content (escaped)
        result.append(new_mif_path)
        # Append from end of ImportObFileDI's content to start of ImportObFile's content
        result.append(mif_str[di_match.end(2) : ob_match.start(2)])
        # Append new ImportObFile content
        result.append(new_path)
        
        # Update pos to end of ImportObFile content
        pos = ob_match.end(2)
        rewrite_count += 1
        log.info(f"  [OK] {basename} -> {new_path} (updated DI and OB)")

    result.append(mif_str[pos:])
    new_mif = "".join(result)

    log.info(
        f"update_xlf_references: rewrote {rewrite_count} <ImportObFileDI>/<ImportObFile> "
        f"reference(s) inside the MIF blob"
    )
    if rewrite_count == 0:
        log.warning("  No rewrites fired — dumping <ImportObFileDI> values found:")
        for i, m in enumerate(_DI_RE.finditer(mif_str)):
            if i >= 10:
                log.warning(f"  …({sum(1 for _ in _DI_RE.finditer(mif_str)) - 10} more)")
                break
            log.warning(f"    [{i}] {m.group(2)!r}")
        log.warning("  Available basenames to match against:")
        for bn in filename_to_new:
            log.warning(f"    {bn!r}")
        return
    if miss_samples:
        log.info(f"  ({len(miss_samples)} <ImportObFileDI> value(s) intentionally left alone — none matched a translated file):")
        for s in miss_samples[:5]:
            log.info(f"    skip: {s!r}")

    raw = new_mif.encode("utf-8", errors="replace")
    if was_gzip:
        raw = gzip.compress(raw)
    internal_el.text = base64.b64encode(raw).decode("ascii")

    tree.write(str(xlf_path), encoding="utf-8", xml_declaration=True)
    log.info(f"update_xlf_references: saved updated XLF → {xlf_path}")


def translate_file(input_path, output_root, target_lang, args, model_to_use, progress_callback=None):
    input_path  = Path(input_path)
    output_root = Path(output_root)

    output_root.mkdir(parents=True, exist_ok=True)
    
    xlf_dir = output_root / "text_conversion_file"
    xlf_dir.mkdir(parents=True, exist_ok=True)
    xlf_out_path = xlf_dir / input_path.name

    tree, root, ns = load_xliff(str(input_path))
    all_units = extract_units(root, ns)
    all_units = merge_units(all_units)

    if not all_units:
        log.warning(f"No segments found in {input_path.name} -- skipping")
        return False

    to_translate, skipped = [], []
    for u in all_units:
        cls = classify(u); u["_class"] = cls
        (skipped if cls == "skip" else to_translate).append(u)
    n_safety = sum(1 for u in to_translate if u["_class"] == "safety")
    log.info(f"translate={len(to_translate)} skip={len(skipped)} safety={n_safety}")

    total_segments = len(all_units)
    if progress_callback:
        progress_callback("Analyzing segments...", 0, 1, {
            "total_segments": total_segments,
            "translated_segments": 0,
            "total_graphics": 0,
            "converted_graphics": 0
        })

    ckpt = str(output_root.parent / f"{input_path.stem}.checkpoint.json")
    all_trans = {}
    if getattr(args, "resume", False):
        all_trans = load_checkpoint(ckpt)
        before = len(to_translate)
        to_translate = [u for u in to_translate if u["id"] not in all_trans]
        log.info(f"Resume: {before-len(to_translate)} done, {len(to_translate)} left")

    sys_p   = build_sys(target_lang)
    batches = [to_translate[i:i+args.batch_size]
               for i in range(0, len(to_translate), args.batch_size)]

    print(f"\n{'─'*62}")
    print(f"  File         : {input_path.name}")
    print(f"  Target       : {LANGUAGES.get(target_lang, target_lang)}")
    print(f"  Model        : {model_to_use}")
    print(f"  Segs         : {len(to_translate)} translate | {len(skipped)} skip")
    print(f"  Mode         : {'DRY RUN' if args.dry_run else 'LIVE'}")
    print(f"  Output root  : {output_root}")
    print(f"  XLF path     : {xlf_out_path}")
    print(f"{'─'*62}\n")

    for i, batch in enumerate(batches, 1):
        log.info(f"Batch {i}/{len(batches)}")
        result = translate_batch(batch, target_lang, sys_p, args.dry_run, model_to_use)
        all_trans.update(result)
        if not args.dry_run:
            save_checkpoint(ckpt, all_trans)
        
        translated_count = len([u for u in all_units if u["id"] in all_trans])
        if progress_callback:
            progress_callback(
                f"Translating segments batch {i}/{len(batches)}...",
                i,
                len(batches),
                {
                    "total_segments": total_segments,
                    "translated_segments": translated_count
                }
            )

        if i < len(batches):
            time.sleep(BATCH_DELAY)

    for u in skipped:
        all_trans[u["id"]] = u["source"]

    if progress_callback:
        progress_callback(
            "Writing translation back to XML...",
            total_segments,
            total_segments,
            {
                "total_segments": total_segments,
                "translated_segments": total_segments
            }
        )

    set_header_lang(root, ns, target_lang)
    n = write_back(all_units, all_trans, ns, target_lang)
    log.info(f"Wrote {n} segments")

    strip_seg_source(root, ns)

    if not validate_xml(tree):
        log.error(f"Output XML is malformed for {input_path.name} -- aborting save")
        return False

    try:
        save_xliff(tree, str(xlf_out_path))
        print(f"\n  ✓  Translated XLF saved: {xlf_out_path}")
    except Exception as e:
        log.error(f"Failed to save XLF: {e}")
        return False

    try:
        print("\n" + "─" * 62)
        print("  Processing referenced graphics (OCR)")
        print("─" * 62)

        path_mapping = process_xlf_references(
            input_path,
            target_lang,
            out_folder=output_root / "graphics",  
            rel_prefix="",  
            rename_with_lang=False,  
            out_xlf_path=xlf_out_path,
            src_graphics_folder=getattr(args, "graphics_source_folder", None),  # Pass uploaded input target 
            progress_callback=progress_callback,
        )

        if path_mapping:
            print("\n  Rewriting <ImportObFile> entries in the translated XLF…")
            update_xlf_references(xlf_out_path, path_mapping)

        print("\n  ✓  Graphics processing complete")

    except Exception as e:
        log.warning(f"OCR image processing failed: {e}")
        print(f"\n  ⚠  OCR processing skipped — Reason: {e}")

    try:
        export_safety_review(
            all_units, all_trans,
            str(output_root.parent / f"{input_path.stem}_safety_review.xlsx"),
        )
        print("\n  ✓  Safety review exported (next to the deliverable, not inside it)")
    except Exception as e:
        log.warning(f"Safety review export failed: {e}")

    try:
        if not args.dry_run and Path(ckpt).exists():
            Path(ckpt).unlink()
    except Exception as e:
        log.warning(f"Checkpoint cleanup failed: {e}")

    print("\n" + "═" * 62)
    print("  TRANSLATION COMPLETED SUCCESSFULLY")
    print("═" * 62)
    print(f"\n  Deliverable folder:")
    print(f"    {output_root}")
    print("═" * 62)

    return True

def run_batch(args, model_to_use, target_lang):
    input_folder  = Path(args.batch_folder).expanduser().resolve()
    output_folder = Path(args.output_folder).expanduser().resolve()
    output_folder.mkdir(parents=True, exist_ok=True)

    xlf_files = sorted(
        list(input_folder.glob("*.xlf")) +
        list(input_folder.glob("*.xliff"))
    )

    if not xlf_files:
        print(f"\n  ✗  No .xlf or .xliff files found in: {input_folder}")
        sys.exit(1)

    print(f"\n{'═'*62}")
    print(f"  BATCH MODE")
    print(f"  Input folder : {input_folder}")
    print(f"  Output root  : {output_folder}")
    print(f"  Files found  : {len(xlf_files)}")
    print(f"  Target lang  : {LANGUAGES.get(target_lang, target_lang)}")
    print(f"{'═'*62}\n")

    success, failed = [], []
    for idx, xlf_path in enumerate(xlf_files, 1):
        print(f"\n[{idx}/{len(xlf_files)}]  Processing: {xlf_path.name}")

        ok = translate_file(xlf_path, output_folder, target_lang, args, model_to_use)
        if ok:
            success.append(xlf_path.name)
        else:
            failed.append(xlf_path.name)

    print(f"\n{'═'*62}")
    print(f"  BATCH COMPLETE")
    print(f"  Success : {len(success)}/{len(xlf_files)}")
    if success:
        for f in success:
            print(f"    ✓  {f}")
    if failed:
        print(f"  Failed  : {len(failed)}")
        for f in failed:
            print(f"    ✗  {f}")
    print(f"\n  Output folder: {output_folder}")
    print(f"{'═'*62}\n")

def run_single(args, model_to_use, target_lang):
    input_path = Path(args.input)
    output_root = (
        Path(args.output) if getattr(args, "output", None)
        else input_path.parent / f"translated_{target_lang}"
    )

    ok = translate_file(input_path, output_root, target_lang, args, model_to_use)
    if not ok:
        sys.exit(1)

def main():
    p = argparse.ArgumentParser(
        prog="translate_xliff_openai.py",
        description="Translate FrameMaker XLIFF exports using the OpenAI API.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )

    p.add_argument("--batch-folder", "-B",
        default=None, metavar="DIR",
        help="Folder containing .xlf files to translate in bulk.")
    p.add_argument("--graphics-source-folder", "-G",
        default=None, metavar="DIR",
        help="User-uploaded folder containing the source graphics files to scan.")
    p.add_argument("--output-folder", "-O",
        default=None, metavar="DIR",
        help="Destination folder for batch output "
             "(default: <batch-folder>/translated_<lang>/).")

    p.add_argument("--target", "-t",
        default=None, choices=list(LANGUAGES.keys()), metavar="LANG",
        help="Target language code. If omitted, an interactive menu appears. "
             "Options: " + ", ".join(LANGUAGES.keys()))
    p.add_argument("--model", "-m",
        default=None,
        help=f"OpenAI model to use (default: {MODEL}).")
    p.add_argument("--batch-size", "-b",
        type=int, default=BATCH_SIZE,
        help=f"Segments per API call (default: {BATCH_SIZE}).")
    p.add_argument("--dry-run",
        action="store_true",
        help="Parse and report without calling the API.")
    p.add_argument("--resume",
        action="store_true",
        help="Skip segments already in the checkpoint file.")
    p.add_argument("--verbose", "-v",
        action="store_true",
        help="Show debug-level log messages.")

    args = p.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    model_to_use = args.model or MODEL

    target_lang = args.target or select_language_interactive()

    # Prompts for the necessary directories interactively if parameters aren't explicitly passed
    if not args.batch_folder:
        batch_folder = ask_path(
            "Enter the folder containing your .xlf files",
            must_exist=True, is_file=False
        )
        args.batch_folder = str(batch_folder)

    if not args.graphics_source_folder:
        graphics_src = ask_path(
            "Enter the path to your uploaded source Graphics folder containing files",
            must_exist=True, is_file=False
        )
        args.graphics_source_folder = str(graphics_src)

    if not args.output_folder:
        default_out = Path(args.batch_folder).parent / f"translated_{target_lang}"
        print(f"\n  Output folder will default to: {default_out}")
        custom = input("  Press Enter to accept, or type a different output path: ").strip()
        args.output_folder = custom if custom else str(default_out)

    run_batch(args, model_to_use, target_lang)

    # ── ZIP the language-rooted folder verbatim ──────────────────────────────
    import zipfile
    output_folder = Path(args.output_folder).expanduser().resolve()
    if output_folder.exists() and any(output_folder.iterdir()):
        zip_path = output_folder.parent / f"{output_folder.name}.zip"
        print(f"\n  Creating deliverable zip: {zip_path.name}")

        count = 0
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for path in sorted(output_folder.rglob("*")):
                if not path.is_file():
                    continue
                rel = path.relative_to(output_folder)
                arcname = f"{output_folder.name}/{rel.as_posix()}"
                zf.write(path, arcname=arcname)
                count += 1

        print(f"  ✓ Zip created: {zip_path}  ({count} file(s))")
    else:
        print("\n  (No output files found — zip skipped)")


if __name__ == "__main__":
    main()